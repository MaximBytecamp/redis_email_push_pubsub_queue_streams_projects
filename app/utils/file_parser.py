import csv
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


def _normalize_header(value: object) -> str:
    return str(value or "").strip().lower()


def parse_recipients_file(path: str) -> Iterable[dict[str, str]]:
    file_path = Path(path)
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        with file_path.open("r", encoding="utf-8-sig", newline="") as fh:
            yield from csv.DictReader(fh)
        return

    if suffix == ".xlsx":
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = [_normalize_header(cell) for cell in next(rows, [])]
        for row in rows:
            yield {headers[index]: "" if value is None else str(value) for index, value in enumerate(row)}
        return

    raise ValueError("Поддерживаются только .csv и .xlsx файлы")
